"""
Genera el Excel de resultados: todas las selecciones y los promedios de
grupo en una sola hoja, con bandas de color para que se entienda de un
vistazo dónde empieza cada sección.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

COLUMNAS = ["archivo", "id", "area_um2", "length_um", "revisar_manualmente", "motivo"]
ANCHOS = [40, 6, 14, 14, 18, 38]

FILL_SELECCION = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
FILL_PROMEDIO = PatternFill("solid", fgColor="C6E0B4")
FILL_GRUPO = PatternFill("solid", fgColor="7030A0")
FILL_GRUPO_PROMEDIO = PatternFill("solid", fgColor="FFD966")

FONT_BANDA = Font(color="FFFFFF", bold=True, size=12)
FONT_BOLD = Font(bold=True)


def generar_excel(selecciones, grupos, path):
    """
    selecciones: lista de dicts {"nombre", "objetivo" (opcional), "filas"
        (lista de dicts con las columnas de COLUMNAS), "promedio_area",
        "promedio_length", "n"}
    grupos: lista de dicts {"nombre", "detalle" (lista de tuplas
        (nombre_seleccion, promedio_area, promedio_length)),
        "promedio_area", "promedio_length"}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mediciones"

    for i, ancho in enumerate(ANCHOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    fila = 1

    def escribir_banda(texto, fill, font):
        nonlocal fila
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(COLUMNAS))
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.fill = fill
        celda.font = font
        celda.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[fila].height = 22
        fila += 1

    def escribir_headers():
        nonlocal fila
        for c, nombre_col in enumerate(COLUMNAS, start=1):
            celda = ws.cell(row=fila, column=c, value=nombre_col)
            celda.fill = FILL_HEADER
            celda.font = FONT_BOLD
        fila += 1

    for sel in selecciones:
        objetivo = sel.get("objetivo")
        texto_banda = f"Selección: {sel['nombre']} (objetivo {objetivo})" if objetivo else f"Selección: {sel['nombre']}"
        escribir_banda(texto_banda, FILL_SELECCION, FONT_BANDA)
        escribir_headers()

        for f in sel["filas"]:
            valores = [f.get("archivo"), f.get("id"), f.get("area_um2"), f.get("length_um"),
                       f.get("revisar_manualmente"), f.get("motivo")]
            for c, v in enumerate(valores, start=1):
                ws.cell(row=fila, column=c, value=v)
            fila += 1

        prom_area = sel["promedio_area"]
        prom_len = sel["promedio_length"]
        etiqueta = f"Promedio de la selección (n={sel['n']} gusanos, sin contar los marcados para revisar)"
        valores_prom = [
            etiqueta, "",
            round(prom_area, 1) if prom_area is not None else "sin datos",
            round(prom_len, 1) if prom_len is not None else "sin datos",
            "", "",
        ]
        for c, v in enumerate(valores_prom, start=1):
            celda = ws.cell(row=fila, column=c, value=v)
            celda.fill = FILL_PROMEDIO
            celda.font = FONT_BOLD
        fila += 2

    if grupos:
        escribir_banda("Promedios por grupo (promedio de los promedios de cada selección)", FILL_GRUPO, FONT_BANDA)
        headers_grupo = ["grupo / selección incluida", "", "area_um2 promedio", "length_um promedio", "", ""]
        for c, v in enumerate(headers_grupo, start=1):
            celda = ws.cell(row=fila, column=c, value=v)
            celda.fill = FILL_HEADER
            celda.font = FONT_BOLD
        fila += 1

        for grupo in grupos:
            for nombre_sel, prom_area_sel, prom_len_sel in grupo["detalle"]:
                valores = [
                    f"    {nombre_sel}", "",
                    round(prom_area_sel, 1) if prom_area_sel is not None else "sin datos",
                    round(prom_len_sel, 1) if prom_len_sel is not None else "sin datos",
                    "", "",
                ]
                for c, v in enumerate(valores, start=1):
                    ws.cell(row=fila, column=c, value=v)
                fila += 1

            valores_prom = [
                f"Promedio del grupo: {grupo['nombre']}", "",
                round(grupo["promedio_area"], 1) if grupo["promedio_area"] is not None else "sin datos",
                round(grupo["promedio_length"], 1) if grupo["promedio_length"] is not None else "sin datos",
                "", "",
            ]
            for c, v in enumerate(valores_prom, start=1):
                celda = ws.cell(row=fila, column=c, value=v)
                celda.fill = FILL_GRUPO_PROMEDIO
                celda.font = FONT_BOLD
            fila += 2

    wb.save(path)
